"""
Rebuild the embedded RAW_DATA block in index.html from the Excel workbook.

The dashboard is intentionally self-contained. This script reads source values
from the workbook with openpyxl, builds the JSON object expected by the existing
JavaScript, and replaces only the `const RAW_DATA = ...;` block in index.html.

Compatible with Python 3.12.
"""

from __future__ import annotations

import json
import re
import sys
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_FILE = SCRIPT_DIR / "CPS FY 25 26 to Fy 2627.xlsx"
HTML_FILE = SCRIPT_DIR / "index.html"

# The project contains inspect.py, which can shadow Python's stdlib inspect
# module when openpyxl imports it. Remove the script directory from import
# lookup before importing third-party libraries.
sys.path = [
    entry
    for entry in sys.path
    if Path(entry or ".").resolve() != SCRIPT_DIR
]

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from dataclasses import dataclass

RAW_DATA_DECLARATION = "const RAW_DATA = "

METRIC_KEYS = {
    "revenue": "revenue",
    "retail count": "retail_count",
    "retail_count": "retail_count",
    "cm2": "cm2",
    "cm2 %": "cm2_pct",
    "cm2%": "cm2_pct",
    "cm2 pct": "cm2_pct",
    "cm2 percentage": "cm2_pct",
}

METRIC_ORDER = ("revenue", "retail_count", "cm2", "cm2_pct")

COLOR_PALETTE = (
    "#E74C3C",
    "#3498DB",
    "#F39C12",
    "#1ABC9C",
    "#9B59B6",
    "#E67E22",
    "#2ECC71",
    "#34495E",
    "#FF6B6B",
    "#00D2D3",
    "#6366F1",
    "#14B8A6",
    "#F43F5E",
    "#84CC16",
    "#A855F7",
    "#0EA5E9",
)


@dataclass(frozen=True)
class Layout:
    """Workbook coordinates discovered from the header rows."""

    header_row: int
    oem_row: int
    fy_col: int
    month_col: int
    metric_columns: OrderedDict[str, dict[str, int]]
    aggregate_columns: dict[str, int]


def main() -> None:
    print("=" * 58)
    print("CPS OEM Dashboard - rebuild from Excel")
    print("=" * 58)

    workbook = load_workbook(EXCEL_FILE, data_only=True)
    worksheet = workbook.active

    layout = detect_layout(worksheet)
    raw_data = build_raw_data(worksheet, layout)
    inject_raw_data(HTML_FILE, raw_data)

    print(f"Workbook: {EXCEL_FILE.name}")
    print(f"Sheet: {worksheet.title}")
    print(f"Months: {len(raw_data['months'])}")
    print(f"FY periods: {', '.join(raw_data['fy_splits'].keys())}")
    print(f"OEM/group series: {len(raw_data['oems'])}")
    print(f"Updated: {HTML_FILE.name}")


def detect_layout(ws: Worksheet) -> Layout:
    """Find FY, month, OEM/group, and metric columns without fixed positions."""

    header_row = find_metric_header_row(ws)
    oem_row = header_row - 1
    if oem_row < 1:
        raise ValueError("Could not find an OEM/group header row above the metric row.")

    fy_col = None
    month_col = None
    for col in range(1, ws.max_column + 1):
        value = normalize_header(ws.cell(header_row, col).value)
        if value in {"fy year", "financial year", "fy"}:
            fy_col = col
        elif value in {"billing month", "month", "billing period"}:
            month_col = col

    if fy_col is None or month_col is None:
        raise ValueError("Could not detect FY Year and Billing Month columns.")

    metric_columns: OrderedDict[str, dict[str, int]] = OrderedDict()
    aggregate_columns: dict[str, int] = {}
    active_group: str | None = None

    for col in range(1, ws.max_column + 1):
        group_value = ws.cell(oem_row, col).value
        if group_value not in (None, ""):
            active_group = clean_label(group_value)

        metric_key = metric_name_to_key(ws.cell(header_row, col).value)
        if not active_group or not metric_key:
            continue

        if is_aggregate_group(active_group):
            aggregate_columns[metric_key] = col
            continue

        metric_columns.setdefault(
            active_group, {key: None for key in METRIC_ORDER}  # type: ignore[dict-item]
        )
        metric_columns[active_group][metric_key] = col

    missing = {
        group: [metric for metric, col in columns.items() if col is None]
        for group, columns in metric_columns.items()
        if any(col is None for col in columns.values())
    }
    if missing:
        details = "; ".join(f"{group}: {', '.join(metrics)}" for group, metrics in missing.items())
        raise ValueError(f"Incomplete metric columns detected: {details}")

    if not metric_columns:
        raise ValueError("No OEM/group metric columns were detected.")

    return Layout(
        header_row=header_row,
        oem_row=oem_row,
        fy_col=fy_col,
        month_col=month_col,
        metric_columns=metric_columns,
        aggregate_columns=aggregate_columns,
    )


def find_metric_header_row(ws: Worksheet) -> int:
    for row in range(1, ws.max_row + 1):
        headers = {normalize_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        metric_hits = {"revenue", "retail count", "cm2", "cm2 %"} & headers
        has_period_headers = bool({"fy year", "billing month"} & headers)
        if len(metric_hits) >= 2 and has_period_headers:
            return row
    raise ValueError("Could not detect the worksheet metric header row.")


def build_raw_data(ws: Worksheet, layout: Layout) -> dict[str, Any]:
    rows = list(iter_data_rows(ws, layout))
    if not rows:
        raise ValueError("No data rows were found below the header.")

    months = [row["month_label"] for row in rows]
    fy_splits = build_fy_splits(rows)
    quarters = build_quarters(rows)
    total_columns = require_aggregate_columns(layout)
    oems = build_oem_series(ws, layout.metric_columns, rows)
    overall_series = build_single_series(ws, total_columns, rows)
    fy_totals = build_fy_totals(ws, layout, rows)
    growth = build_growth(fy_totals)

    return {
        "months": months,
        "fy_splits": fy_splits,
        "quarters": quarters,
        "fy_totals": fy_totals,
        "growth": growth,
        "oems": oems,
        "overall_series": overall_series,
        "colors": build_colors(oems.keys()),
    }


def iter_data_rows(ws: Worksheet, layout: Layout) -> Iterable[dict[str, Any]]:
    for row in range(layout.header_row + 1, ws.max_row + 1):
        fy_value = ws.cell(row, layout.fy_col).value
        month_value = ws.cell(row, layout.month_col).value
        if is_blank(fy_value) and is_blank(month_value):
            continue
        if is_blank(month_value):
            continue

        yield {
            "row": row,
            "fy_label": clean_label(fy_value),
            "fy_key": fy_to_key(fy_value),
            "month_label": month_to_label(month_value),
            "month_value": month_value,
        }


def build_oem_series(
    ws: Worksheet,
    metric_columns: OrderedDict[str, dict[str, int]],
    rows: list[dict[str, Any]],
) -> OrderedDict[str, dict[str, list[Any]]]:
    oems: OrderedDict[str, dict[str, list[Any]]] = OrderedDict()

    for group, columns in metric_columns.items():
        oems[group] = {metric: [] for metric in METRIC_ORDER}
        for row_info in rows:
            row_number = row_info["row"]
            for metric in METRIC_ORDER:
                col = columns[metric]
                oems[group][metric].append(cell_to_json_value(ws.cell(row_number, col).value))

    return oems


def build_single_series(
    ws: Worksheet, columns: dict[str, int], rows: list[dict[str, Any]]
) -> dict[str, list[Any]]:
    series = {metric: [] for metric in METRIC_ORDER}
    for row_info in rows:
        row_number = row_info["row"]
        for metric in METRIC_ORDER:
            series[metric].append(cell_to_json_value(ws.cell(row_number, columns[metric]).value))
    return series


def build_fy_splits(rows: list[dict[str, Any]]) -> OrderedDict[str, list[str]]:
    fy_splits: OrderedDict[str, list[str]] = OrderedDict()
    fy_splits["overall"] = [row["month_label"] for row in rows]

    for row in rows:
        fy_splits.setdefault(row["fy_key"], []).append(row["month_label"])

    return fy_splits


def build_quarters(rows: list[dict[str, Any]]) -> OrderedDict[str, dict[str, list[str]]]:
    quarters: OrderedDict[str, dict[str, list[str]]] = OrderedDict()
    for fy_key, fy_rows in group_rows_by_fy(rows).items():
        quarters[fy_key] = {}
        for offset, row in enumerate(fy_rows):
            quarter = f"Q{offset // 3 + 1}"
            quarters[fy_key].setdefault(quarter, []).append(row["month_label"])
    return quarters


def build_fy_totals(
    ws: Worksheet, layout: Layout, rows: list[dict[str, Any]]
) -> OrderedDict[str, dict[str, Any]]:
    source_columns = require_aggregate_columns(layout)
    totals: OrderedDict[str, dict[str, Any]] = OrderedDict()

    for key, grouped_rows in [("overall", rows), *group_rows_by_fy(rows).items()]:
        totals[key] = {
            "revenue": sum_metric(ws, grouped_rows, source_columns["revenue"]),
            "cm2": sum_metric(ws, grouped_rows, source_columns["cm2"]),
            "retail_count": sum_metric(ws, grouped_rows, source_columns["retail_count"]),
            "cm2_pct": ratio_or_none(
                sum_metric(ws, grouped_rows, source_columns["cm2"]),
                sum_metric(ws, grouped_rows, source_columns["revenue"]),
            ),
        }

    return totals


def require_aggregate_columns(layout: Layout) -> dict[str, int]:
    if not all(metric in layout.aggregate_columns for metric in METRIC_ORDER):
        raise ValueError(
            "Could not detect a complete aggregate group such as Overall. "
            "Totals and growth are read from aggregate columns by design."
        )
    return layout.aggregate_columns


def build_growth(fy_totals: OrderedDict[str, dict[str, Any]]) -> dict[str, float]:
    fy_keys = [key for key in fy_totals if key != "overall"]
    if len(fy_keys) < 2:
        return {"revenue_growth_pct": 0.0, "retail_growth_pct": 0.0, "cm2_growth_pct": 0.0}

    baseline = fy_totals[fy_keys[0]]
    current = fy_totals[fy_keys[-1]]
    return {
        "revenue_growth_pct": growth_pct(current["revenue"], baseline["revenue"]),
        "retail_growth_pct": growth_pct(current["retail_count"], baseline["retail_count"]),
        "cm2_growth_pct": growth_pct(current["cm2"], baseline["cm2"]),
    }


def group_rows_by_fy(rows: list[dict[str, Any]]) -> OrderedDict[str, list[dict[str, Any]]]:
    grouped: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for row in rows:
        grouped.setdefault(row["fy_key"], []).append(row)
    return grouped


def sum_metric(ws: Worksheet, rows: list[dict[str, Any]], column: int) -> float:
    total = 0.0
    for row in rows:
        total += numeric_value(ws.cell(row["row"], column).value)
    return total


def numeric_value(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def ratio_or_none(numerator: float, denominator: float) -> float | None:
    if denominator == 0:
        return None
    return numerator / denominator


def growth_pct(current: Any, baseline: Any) -> float:
    if not isinstance(current, (int, float)) or not isinstance(baseline, (int, float)) or baseline == 0:
        return 0.0
    return round(((current - baseline) / baseline) * 100, 2)


def build_colors(names: Iterable[str]) -> dict[str, str]:
    return {name: COLOR_PALETTE[index % len(COLOR_PALETTE)] for index, name in enumerate(names)}


def inject_raw_data(html_path: Path, raw_data: dict[str, Any]) -> None:
    html = html_path.read_text(encoding="utf-8")
    start = html.find(RAW_DATA_DECLARATION)
    if start == -1:
        raise ValueError(f"Could not find `{RAW_DATA_DECLARATION.strip()}` in {html_path.name}.")

    object_start = html.find("{", start)
    if object_start == -1:
        raise ValueError("Could not find the start of the RAW_DATA object.")

    object_end = find_matching_brace(html, object_start)
    semicolon = html.find(";", object_end)
    if semicolon == -1:
        raise ValueError("Could not find the semicolon after the RAW_DATA object.")

    json_text = json.dumps(raw_data, indent=2, ensure_ascii=False)
    replacement = f"{RAW_DATA_DECLARATION}{json_text};"
    updated = html[:start] + replacement + html[semicolon + 1 :]

    html_path.write_text(updated, encoding="utf-8")


def find_matching_brace(text: str, open_index: int) -> int:
    depth = 0
    in_string = False
    escape = False

    for index in range(open_index, len(text)):
        char = text[index]

        if in_string:
            if escape:
                escape = False
            elif char == "\\":
                escape = True
            elif char == '"':
                in_string = False
            continue

        if char == '"':
            in_string = True
        elif char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return index

    raise ValueError("RAW_DATA object braces are not balanced.")


def metric_name_to_key(value: Any) -> str | None:
    return METRIC_KEYS.get(normalize_header(value))


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def clean_label(value: Any) -> str:
    label = re.sub(r"\s+", " ", str(value or "").strip())
    label = re.sub(r"\(\s+", "(", label)
    label = re.sub(r"\s+\)", ")", label)
    return label


def is_aggregate_group(label: str) -> bool:
    return normalize_header(label) in {"overall", "total", "grand total"}


def fy_to_key(value: Any) -> str:
    label = clean_label(value)
    digits = re.findall(r"\d+", label)
    if len(digits) >= 2:
        return "fy" + "".join(digits[:2])
    return re.sub(r"[^a-z0-9]+", "", normalize_header(label)) or "unknown"


def month_to_label(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%b-%Y")
    if isinstance(value, date):
        return value.strftime("%b-%Y")
    return clean_label(value)


def cell_to_json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


if __name__ == "__main__":
    main()
