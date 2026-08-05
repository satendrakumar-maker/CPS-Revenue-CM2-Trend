from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook(r'CPS FY 25 26 to Fy 2627.xlsx')
ws = wb.active

print('Sheets:', wb.sheetnames)
print(f'Active: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}')
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f'Row {i}:', row)
    if i > 18:
        break
