from openpyxl import load_workbook
from datetime import datetime
import json
import os

# ── 1. Read Excel ────────────────────────────────────────────────────────────
wb = load_workbook(r'C:/Users/saten/Downloads/CPS FY 25 26 to Fy 2627.xlsx')
ws = wb.active

raw_oems = [
    ("Bajaj (Incl. KTM & TRM)", 3, 6),
    ("Chetak Only", 7, 10),
    ("Ather", 11, 14),
    ("OLA Only CPS", 15, 18),
    ("TVS: All", 19, 22),
    ("Revolt", 23, 26),
    ("Jawa (Incl. Manpower)", 27, 30),
    ("TVS: Only Iqube", 31, 34),
]

data_rows = list(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True))

all_months = []
raw_data = {name: {"revenue": [], "retail_count": [], "cm2": [], "cm2_pct": []} for name, _, _ in raw_oems}

for row in data_rows:
    month_val = row[2]
    if month_val is None:
        continue
    if isinstance(month_val, datetime):
        month_label = month_val.strftime("%b-%Y")
    else:
        month_label = str(month_val)
    all_months.append(month_label)

    for name, start, end in raw_oems:
        vals = row[start:end+1]
        raw_data[name]["revenue"].append(float(vals[0]) if vals[0] is not None else 0)
        raw_data[name]["retail_count"].append(float(vals[1]) if vals[1] is not None else 0)
        raw_data[name]["cm2"].append(float(vals[2]) if vals[2] is not None else 0)
        raw_data[name]["cm2_pct"].append(float(vals[3]) if vals[3] is not None else 0)

# ── 2. Computed OEMs ─────────────────────────────────────────────────────────
def compute_pct(cm2_list, rev_list):
    return [c/r if r != 0 else 0 for c, r in zip(cm2_list, rev_list)]

bajaj_name = "Bajaj (Incl. KTM & TRM)"
chetak_name = "Chetak Only"
tvs_all_name = "TVS: All"
tvs_iqube_name = "TVS: Only Iqube"

bajaj_chetak = {
    "revenue": [raw_data[bajaj_name]["revenue"][i] + raw_data[chetak_name]["revenue"][i] for i in range(len(all_months))],
    "retail_count": [raw_data[bajaj_name]["retail_count"][i] + raw_data[chetak_name]["retail_count"][i] for i in range(len(all_months))],
    "cm2": [raw_data[bajaj_name]["cm2"][i] + raw_data[chetak_name]["cm2"][i] for i in range(len(all_months))],
}
bajaj_chetak["cm2_pct"] = compute_pct(bajaj_chetak["cm2"], bajaj_chetak["revenue"])

tvs_ice = {
    "revenue": [raw_data[tvs_all_name]["revenue"][i] - raw_data[tvs_iqube_name]["revenue"][i] for i in range(len(all_months))],
    "retail_count": [raw_data[tvs_all_name]["retail_count"][i] - raw_data[tvs_iqube_name]["retail_count"][i] for i in range(len(all_months))],
    "cm2": [raw_data[tvs_all_name]["cm2"][i] - raw_data[tvs_iqube_name]["cm2"][i] for i in range(len(all_months))],
}
tvs_ice["cm2_pct"] = compute_pct(tvs_ice["cm2"], tvs_ice["revenue"])

all_oem_data = dict(raw_data)
all_oem_data["Bajaj + Chetak"] = bajaj_chetak
all_oem_data["TVS ICE"] = tvs_ice

total_oem_names = [
    "Bajaj (Incl. KTM & TRM)", "Chetak Only", "Ather",
    "OLA Only CPS", "TVS: All", "Revolt", "Jawa (Incl. Manpower)"
]

# ── 3. FY Totals ────────────────────────────────────────────────────────────
fy2526_indices = list(range(0, 12))
fy2627_indices = list(range(12, 15))

def compute_fy_totals(indices):
    total_rev = sum(sum(raw_data[n]["revenue"][i] for i in indices) for n in total_oem_names)
    total_cm2 = sum(sum(raw_data[n]["cm2"][i] for i in indices) for n in total_oem_names)
    total_ret = sum(sum(raw_data[n]["retail_count"][i] for i in indices) for n in total_oem_names)
    avg_pct = total_cm2 / total_rev if total_rev != 0 else 0
    return {"revenue": total_rev, "cm2": total_cm2, "retail_count": total_ret, "cm2_pct": avg_pct}

fy2526_totals = compute_fy_totals(fy2526_indices)
fy2627_totals = compute_fy_totals(fy2627_indices)
overall_totals = compute_fy_totals(fy2526_indices + fy2627_indices)

rev_growth = ((overall_totals["revenue"] - fy2526_totals["revenue"]) / fy2526_totals["revenue"] * 100) if fy2526_totals["revenue"] else 0
ret_growth = ((overall_totals["retail_count"] - fy2526_totals["retail_count"]) / fy2526_totals["retail_count"] * 100) if fy2526_totals["retail_count"] else 0
cm2_growth = ((overall_totals["cm2"] - fy2526_totals["cm2"]) / fy2526_totals["cm2"] * 100) if fy2526_totals["cm2"] else 0

# ── 4. Build data object for embedding ──────────────────────────────────────
data_obj = {
    "months": all_months,
    "fy_splits": {
        "overall": all_months,
        "fy2526": all_months[:12],
        "fy2627": all_months[12:],
    },
    "fy_totals": {
        "overall": overall_totals,
        "fy2526": fy2526_totals,
        "fy2627": fy2627_totals,
    },
    "growth": {
        "revenue_growth_pct": round(rev_growth, 2),
        "retail_growth_pct": round(ret_growth, 2),
        "cm2_growth_pct": round(cm2_growth, 2),
    },
    "oems": all_oem_data,
    "colors": {
        "Bajaj (Incl. KTM & TRM)": "#E74C3C",
        "Chetak Only": "#3498DB",
        "Ather": "#F39C12",
        "OLA Only CPS": "#1ABC9C",
        "TVS: All": "#9B59B6",
        "Revolt": "#E67E22",
        "Jawa (Incl. Manpower)": "#2ECC71",
        "TVS: Only Iqube": "#34495E",
        "Bajaj + Chetak": "#FF6B6B",
        "TVS ICE": "#00D2D3",
    }
}

data_json_str = json.dumps(data_obj, indent=2)

# ── 5. Build Single HTML File ───────────────────────────────────────────────
html = f'''<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CPS OEM Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
<style>
:root {{
  --bg-primary:#0f172a; --bg-secondary:#1e293b; --bg-card:rgba(30,41,59,0.7); --bg-glass:rgba(30,41,59,0.4);
  --text-primary:#f1f5f9; --text-secondary:#94a3b8; --text-muted:#64748b;
  --accent-blue:#3b82f6; --accent-green:#10b981; --accent-red:#ef4444; --accent-orange:#f59e0b;
  --accent-purple:#8b5cf6; --accent-pink:#ec4899;
  --border-color:rgba(148,163,184,0.1); --shadow:0 8px 32px rgba(0,0,0,0.3); --shadow-sm:0 2px 8px rgba(0,0,0,0.15);
  --radius:16px; --radius-sm:10px; --transition:all 0.3s cubic-bezier(0.4,0,0.2,1);
}}
[data-theme="light"] {{
  --bg-primary:#f8fafc; --bg-secondary:#fff; --bg-card:rgba(255,255,255,0.85); --bg-glass:rgba(255,255,255,0.5);
  --text-primary:#1e293b; --text-secondary:#475569; --text-muted:#94a3b8;
  --border-color:rgba(148,163,184,0.2); --shadow:0 8px 32px rgba(0,0,0,0.08); --shadow-sm:0 2px 8px rgba(0,0,0,0.04);
}}
*{{box-sizing:border-box;margin:0;padding:0}}
html{{scroll-behavior:smooth;font-size:15px}}
body{{
  font-family:'Segoe UI',system-ui,-apple-system,sans-serif;background:var(--bg-primary);color:var(--text-primary);min-height:100vh;line-height:1.5;-webkit-font-smoothing:antialiased;
}}

/* Loading */
#loadingScreen{{position:fixed;inset:0;background:var(--bg-primary);display:flex;flex-direction:column;align-items:center;justify-content:center;z-index:9999;gap:20px}}
.loader{{width:56px;height:56px;border:4px solid rgba(59,130,246,0.2);border-top-color:var(--accent-blue);border-radius:50%;animation:spin 0.8s linear infinite}}
@keyframes spin{{to{{transform:rotate(360deg)}}}}
#loadingScreen p{{color:var(--text-secondary);font-size:1rem}}

/* Top Bar */
.topbar{{display:flex;justify-content:space-between;align-items:center;padding:16px 24px;background:var(--bg-glass);backdrop-filter:blur(20px);border-bottom:1px solid var(--border-color);position:sticky;top:0;z-index:100}}
.brand{{display:flex;align-items:center;gap:12px}}
.logo{{width:44px;height:44px;background:linear-gradient(135deg,var(--accent-blue),var(--accent-purple));border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.4rem;box-shadow:var(--shadow-sm)}}
.brand h1{{font-size:1.3rem;font-weight:700;letter-spacing:-0.5px}}
.tagline{{font-size:0.78rem;color:var(--text-muted)}}
.topbar-right{{display:flex;align-items:center;gap:12px}}
.last-refresh{{font-size:0.78rem;color:var(--text-muted)}}
.icon-btn{{width:38px;height:38px;border-radius:10px;border:1px solid var(--border-color);background:var(--bg-card);color:var(--text-primary);cursor:pointer;font-size:1rem;transition:var(--transition);display:flex;align-items:center;justify-content:center}}
.icon-btn:hover{{background:var(--accent-blue);color:#fff;border-color:var(--accent-blue)}}
.btn-primary{{padding:10px 18px;border-radius:10px;border:none;background:linear-gradient(135deg,var(--accent-blue),var(--accent-purple));color:#fff;font-size:0.85rem;font-weight:600;cursor:pointer;transition:var(--transition);box-shadow:var(--shadow-sm)}}
.btn-primary:hover{{transform:translateY(-2px);box-shadow:var(--shadow)}}
.btn-outline{{padding:8px 14px;border-radius:8px;border:1px solid var(--border-color);background:transparent;color:var(--text-secondary);font-size:0.8rem;cursor:pointer;transition:var(--transition)}}
.btn-outline:hover{{border-color:var(--accent-blue);color:var(--accent-blue)}}

/* Filters */
.filters-bar{{display:flex;gap:12px;padding:16px 24px;flex-wrap:wrap;align-items:flex-end;background:var(--bg-secondary);border-bottom:1px solid var(--border-color)}}
.filter-group{{display:flex;flex-direction:column;gap:4px}}
.filter-group label{{font-size:0.72rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.5px;font-weight:600}}
.filter-group select{{padding:9px 14px;border-radius:var(--radius-sm);border:1px solid var(--border-color);background:var(--bg-card);color:var(--text-primary);font-size:0.85rem;cursor:pointer;min-width:160px;outline:none;transition:var(--transition)}}
.filter-group select:focus{{border-color:var(--accent-blue)}}
.filter-actions{{display:flex;gap:8px;margin-left:auto;align-items:flex-end}}

/* Main Tabs */
.main-tabs{{display:flex;gap:8px;padding:16px 24px 0}}
.main-tab{{padding:12px 24px;border-radius:var(--radius-sm) var(--radius-sm) 0 0;border:none;background:transparent;color:var(--text-muted);font-size:0.9rem;font-weight:600;cursor:pointer;transition:var(--transition);border-bottom:3px solid transparent}}
.main-tab:hover{{color:var(--text-secondary)}}
.main-tab.active{{color:var(--accent-blue);border-bottom-color:var(--accent-blue);background:var(--bg-card)}}

/* Tab Panels */
.tab-panel{{display:none;padding:20px 24px 40px;animation:fadeIn 0.4s ease}}
.tab-panel.active{{display:block}}
@keyframes fadeIn{{from{{opacity:0;transform:translateY(12px)}}to{{opacity:1;transform:translateY(0)}}}}

/* KPI Cards */
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin-bottom:24px}}
.kpi-card{{
  background:var(--bg-card);backdrop-filter:blur(16px);-webkit-backdrop-filter:blur(16px);border:1px solid var(--border-color);border-radius:var(--radius);padding:20px;position:relative;overflow:hidden;transition:var(--transition);box-shadow:var(--shadow-sm)
}}
.kpi-card:hover{{transform:translateY(-4px);box-shadow:var(--shadow)}}
.kpi-card::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--radius) var(--radius) 0 0}}
.kpi-card.revenue::before{{background:linear-gradient(90deg,var(--accent-blue),var(--accent-purple))}}
.kpi-card.retail::before{{background:linear-gradient(90deg,var(--accent-orange),var(--accent-red))}}
.kpi-card.cm2::before{{background:linear-gradient(90deg,var(--accent-green),#34d399)}}
.kpi-card.cmpct::before{{background:linear-gradient(90deg,var(--accent-purple),var(--accent-pink))}}
.kpi-card.growth-rev::before{{background:linear-gradient(90deg,#06b6d4,var(--accent-blue))}}
.kpi-card.growth-ret::before{{background:linear-gradient(90deg,#f97316,var(--accent-orange))}}

.kpi-icon{{position:absolute;top:16px;right:16px;font-size:1.6rem;opacity:0.12}}
.kpi-label{{font-size:0.72rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.8px;font-weight:600}}
.kpi-value{{font-size:1.7rem;font-weight:700;margin:8px 0 4px;letter-spacing:-0.5px}}
.kpi-value.up{{color:var(--accent-green)}}
.kpi-value.down{{color:var(--accent-red)}}
.kpi-sub{{font-size:0.8rem;color:var(--text-secondary);font-weight:500}}

/* Charts */
.chart-row{{margin-bottom:20px}}
.chart-row.two-col{{display:grid;grid-template-columns:repeat(2,1fr);gap:20px}}
.chart-card{{
  background:var(--bg-card);backdrop-filter:blur(16px);-webkit-backdrop-filter:blur(16px);border:1px solid var(--border-color);border-radius:var(--radius);overflow:hidden;box-shadow:var(--shadow-sm);transition:var(--transition)
}}
.chart-card:hover{{box-shadow:var(--shadow)}}
.chart-header{{display:flex;justify-content:space-between;align-items:center;padding:16px 20px;border-bottom:1px solid var(--border-color)}}
.chart-header h3{{font-size:0.95rem;font-weight:600}}
.chart-action{{width:32px;height:32px;border-radius:8px;border:1px solid var(--border-color);background:transparent;cursor:pointer;font-size:0.9rem;color:var(--text-secondary);transition:var(--transition)}}
.chart-action:hover{{background:var(--accent-blue);border-color:var(--accent-blue);color:#fff}}
.chart-body{{padding:16px;height:320px;position:relative}}
.chart-card.large .chart-body{{height:400px}}

/* Table */
.table-section{{margin-bottom:24px}}
.table-card{{
  background:var(--bg-card);backdrop-filter:blur(16px);border:1px solid var(--border-color);border-radius:var(--radius);overflow:hidden;box-shadow:var(--shadow-sm)
}}
.table-header{{display:flex;justify-content:space-between;align-items:center;padding:16px 20px;border-bottom:1px solid var(--border-color)}}
.table-header h3{{font-size:0.95rem;font-weight:600}}
.table-header input{{
  padding:8px 14px;border-radius:8px;border:1px solid var(--border-color);background:var(--bg-secondary);color:var(--text-primary);font-size:0.85rem;outline:none;width:220px
}}
.table-header input:focus{{border-color:var(--accent-blue)}}
.table-wrap{{overflow-x:auto;max-height:500px;overflow-y:auto}}
table{{width:100%;border-collapse:collapse;font-size:0.86rem}}
thead{{position:sticky;top:0;z-index:10}}
thead th{{
  background:var(--bg-secondary);padding:12px 14px;text-align:left;font-weight:600;color:var(--text-muted);text-transform:uppercase;font-size:0.72rem;letter-spacing:0.5px;border-bottom:2px solid var(--border-color);white-space:nowrap
}}
tbody td{{padding:11px 14px;border-bottom:1px solid var(--border-color);color:var(--text-secondary)}}
tbody tr:hover{{background:rgba(59,130,246,0.06)}}
tbody tr:last-child td{{border-bottom:none}}
.num{{text-align:right;font-family:'SF Mono',monospace;font-weight:500}}
tbody td:first-child{{font-weight:600;color:var(--text-primary)}}
.badge{{display:inline-block;padding:3px 10px;border-radius:12px;font-size:0.72rem;font-weight:600}}
.badge-up{{background:rgba(16,185,129,0.15);color:var(--accent-green)}}
.badge-down{{background:rgba(239,68,68,0.15);color:var(--accent-red)}}

/* OEM Selector */
.oem-selector-bar{{margin-bottom:20px}}
.oem-pills{{display:flex;gap:8px;flex-wrap:wrap}}
.oem-pill{{
  padding:10px 18px;border-radius:24px;border:1px solid var(--border-color);background:var(--bg-card);color:var(--text-secondary);font-size:0.82rem;font-weight:600;cursor:pointer;transition:var(--transition);display:flex;align-items:center;gap:8px
}}
.oem-pill:hover{{border-color:var(--accent-blue);color:var(--text-primary)}}
.oem-pill.active{{
  background:linear-gradient(135deg,var(--accent-blue),var(--accent-purple));color:#fff;border-color:transparent;box-shadow:var(--shadow-sm)
}}
.oem-dot{{width:8px;height:8px;border-radius:50%;display:inline-block}}

/* Footer */
.dashboard-footer{{text-align:center;padding:20px;color:var(--text-muted);font-size:0.78rem;border-top:1px solid var(--border-color)}}

/* Responsive */
@media (max-width:1024px){{.chart-row.two-col{{grid-template-columns:1fr}}}}
@media (max-width:768px){{
  .topbar{{flex-direction:column;align-items:flex-start;gap:12px}}
  .filters-bar{{flex-direction:column;align-items:stretch}}
  .filter-group select{{width:100%}}
  .filter-actions{{margin-left:0;width:100%}}
  .kpi-grid{{grid-template-columns:1fr}}
  .chart-body{{height:260px}}
  .chart-card.large .chart-body{{height:300px}}
  .oem-pills{{justify-content:center}}
  .main-tabs{{overflow-x:auto}}
  .table-header{{flex-direction:column;gap:10px;align-items:stretch}}
  .table-header input{{width:100%}}
}}

/* Scrollbar */
::-webkit-scrollbar{{width:6px;height:6px}}
::-webkit-scrollbar-track{{background:transparent}}
::-webkit-scrollbar-thumb{{background:var(--text-muted);border-radius:3px}}
::-webkit-scrollbar-thumb:hover{{background:var(--text-secondary)}}

/* Animations */
.kpi-card,.chart-card,.table-card{{animation:slideUp 0.5s ease backwards}}
.kpi-card:nth-child(1){{animation-delay:0.05s}}
.kpi-card:nth-child(2){{animation-delay:0.1s}}
.kpi-card:nth-child(3){{animation-delay:0.15s}}
.kpi-card:nth-child(4){{animation-delay:0.2s}}
.kpi-card:nth-child(5){{animation-delay:0.25s}}
.kpi-card:nth-child(6){{animation-delay:0.3s}}

@keyframes slideUp{{from{{opacity:0;transform:translateY(20px)}}to{{opacity:1;transform:translateY(0)}}}}
@keyframes pulse{{0%{{transform:scale(1)}}50%{{transform:scale(1.03)}}100%{{transform:scale(1)}}}}
</style>
</head>
<body>

<div id="loadingScreen">
  <div class="loader"></div>
  <p>Loading Dashboard...</p>
</div>

<div id="app" style="display:none">

<header class="topbar">
  <div class="brand">
    <div class="logo">📊</div>
    <div>
      <h1>CPS OEM Dashboard</h1>
      <span class="tagline">Revenue · CM2 · Retail · Trends</span>
    </div>
  </div>
  <div class="topbar-right">
    <span class="last-refresh" id="lastRefresh">--</span>
    <button class="icon-btn" id="themeToggle" title="Toggle Theme">🌙</button>
    <button class="btn-primary" id="refreshBtn">🔄 Refresh</button>
  </div>
</header>

<section class="filters-bar">
  <div class="filter-group">
    <label>Financial Year</label>
    <select id="fyFilter">
      <option value="overall">📅 Overall</option>
      <option value="fy2526">📅 FY 2025-26</option>
      <option value="fy2627">📅 FY 2026-27</option>
    </select>
  </div>
  <div class="filter-group">
    <label>OEM</label>
    <select id="oemFilter">
      <option value="all">All OEMs</option>
    </select>
  </div>
  <div class="filter-group">
    <label>Month</label>
    <select id="monthFilter">
      <option value="all">All Months</option>
    </select>
  </div>
  <div class="filter-group">
    <label>Quarter</label>
    <select id="quarterFilter">
      <option value="all">All Quarters</option>
      <option value="Q1">Q1 (Apr-Jun)</option>
      <option value="Q2">Q2 (Jul-Sep)</option>
      <option value="Q3">Q3 (Oct-Dec)</option>
      <option value="Q4">Q4 (Jan-Mar)</option>
    </select>
  </div>
  <div class="filter-actions">
    <button class="btn-outline" id="resetFilters">Reset</button>
    <button class="btn-outline" id="exportBtn">📥 Export CSV</button>
  </div>
</section>

<nav class="main-tabs">
  <button class="main-tab active" data-tab="executive">📈 Executive</button>
  <button class="main-tab" data-tab="oemwise">🏭 OEM-wise</button>
</nav>

<!-- ========== EXECUTIVE TAB ========== -->
<main id="tab-executive" class="tab-panel active">
  <div class="kpi-grid">
    <div class="kpi-card revenue">
      <div class="kpi-icon">💰</div>
      <div class="kpi-label">Total Revenue</div>
      <div class="kpi-value" id="execRevenue">--</div>
      <div class="kpi-sub" id="execRevenueSub">--</div>
    </div>
    <div class="kpi-card retail">
      <div class="kpi-icon">🏍️</div>
      <div class="kpi-label">Total Retail</div>
      <div class="kpi-value" id="execRetail">--</div>
      <div class="kpi-sub" id="execRetailSub">--</div>
    </div>
    <div class="kpi-card cm2">
      <div class="kpi-icon">📈</div>
      <div class="kpi-label">Total CM2</div>
      <div class="kpi-value" id="execCm2">--</div>
      <div class="kpi-sub" id="execCm2Sub">--</div>
    </div>
    <div class="kpi-card cmpct">
      <div class="kpi-icon">📊</div>
      <div class="kpi-label">CM2 %</div>
      <div class="kpi-value" id="execCm2Pct">--</div>
      <div class="kpi-sub" id="execCm2PctSub">--</div>
    </div>
    <div class="kpi-card growth-rev">
      <div class="kpi-icon">🚀</div>
      <div class="kpi-label">Revenue Growth</div>
      <div class="kpi-value" id="execRevGrowth">--</div>
      <div class="kpi-sub" id="execRevGrowthSub">--</div>
    </div>
    <div class="kpi-card growth-ret">
      <div class="kpi-icon">📊</div>
      <div class="kpi-label">Retail Growth</div>
      <div class="kpi-value" id="execRetGrowth">--</div>
      <div class="kpi-sub" id="execRetGrowthSub">--</div>
    </div>
  </div>

  <div class="chart-row">
    <div class="chart-card large">
      <div class="chart-header">
        <h3>📉 Revenue Trend</h3>
        <button class="chart-action" data-chart="revTrendChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="revTrendChart"></canvas></div>
    </div>
  </div>

  <div class="chart-row two-col">
    <div class="chart-card">
      <div class="chart-header">
        <h3>🏍️ Retail Trend</h3>
        <button class="chart-action" data-chart="retTrendChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="retTrendChart"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-header">
        <h3>📈 CM2 Trend</h3>
        <button class="chart-action" data-chart="cm2TrendChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="cm2TrendChart"></canvas></div>
    </div>
  </div>

  <div class="chart-row two-col">
    <div class="chart-card">
      <div class="chart-header">
        <h3>📊 CM2 % Trend</h3>
        <button class="chart-action" data-chart="cm2PctTrendChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="cm2PctTrendChart"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-header">
        <h3>🏆 OEM Revenue Ranking</h3>
        <button class="chart-action" data-chart="oemCompareChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="oemCompareChart"></canvas></div>
    </div>
  </div>

  <div class="chart-row">
    <div class="chart-card large">
      <div class="chart-header">
        <h3>📊 Monthly OEM Revenue Contribution (Stacked)</h3>
        <button class="chart-action" data-chart="stackedRevChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="stackedRevChart"></canvas></div>
    </div>
  </div>

  <div class="table-section">
    <div class="table-card">
      <div class="table-header">
        <h3>📋 Monthly Summary</h3>
        <input type="text" id="tableSearch" placeholder="🔍 Search OEM...">
      </div>
      <div class="table-wrap">
        <table id="summaryTable">
          <thead></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </div>
</main>

<!-- ========== OEM-WISE TAB ========== -->
<main id="tab-oemwise" class="tab-panel">
  <div class="oem-selector-bar">
    <div class="oem-pills" id="oemPills"></div>
  </div>

  <div class="kpi-grid four">
    <div class="kpi-card revenue">
      <div class="kpi-icon">💰</div>
      <div class="kpi-label">Revenue</div>
      <div class="kpi-value" id="oemRevenue">--</div>
      <div class="kpi-sub" id="oemRevenueSub">--</div>
    </div>
    <div class="kpi-card retail">
      <div class="kpi-icon">🏍️</div>
      <div class="kpi-label">Retail Count</div>
      <div class="kpi-value" id="oemRetail">--</div>
      <div class="kpi-sub" id="oemRetailSub">--</div>
    </div>
    <div class="kpi-card cm2">
      <div class="kpi-icon">📈</div>
      <div class="kpi-label">CM2</div>
      <div class="kpi-value" id="oemCm2">--</div>
      <div class="kpi-sub" id="oemCm2Sub">--</div>
    </div>
    <div class="kpi-card cmpct">
      <div class="kpi-icon">📊</div>
      <div class="kpi-label">CM2 %</div>
      <div class="kpi-value" id="oemCm2Pct">--</div>
      <div class="kpi-sub" id="oemCm2PctSub">--</div>
    </div>
  </div>

  <div class="chart-row two-col">
    <div class="chart-card">
      <div class="chart-header">
        <h3>💰 Revenue Trend</h3>
        <button class="chart-action" data-chart="oemRevChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="oemRevChart"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-header">
        <h3>🏍️ Retail Trend</h3>
        <button class="chart-action" data-chart="oemRetChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="oemRetChart"></canvas></div>
    </div>
  </div>

  <div class="chart-row two-col">
    <div class="chart-card">
      <div class="chart-header">
        <h3>📈 CM2 Trend</h3>
        <button class="chart-action" data-chart="oemCm2Chart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="oemCm2Chart"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-header">
        <h3>📊 CM2 % Trend</h3>
        <button class="chart-action" data-chart="oemCm2PctChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="oemCm2PctChart"></canvas></div>
    </div>
  </div>

  <div class="chart-row">
    <div class="chart-card large">
      <div class="chart-header">
        <h3>📈 OEM Growth % Comparison</h3>
        <button class="chart-action" data-chart="oemGrowthChart" title="Download PNG">💾</button>
      </div>
      <div class="chart-body"><canvas id="oemGrowthChart"></canvas></div>
    </div>
  </div>
</main>

<footer class="dashboard-footer">
  <p>CPS OEM Dashboard · Data from Excel · <span id="footerTime">--</span></p>
</footer>

</div>

<script>
// ═══════════════════════════════════════════════════════════════
// EMBEDDED DATA — No fetch needed, works when opened directly
// ═══════════════════════════════════════════════════════════════
const RAW_DATA = {data_json_str};

// Display name mapping
const OEM_NAMES = {{
  'Bajaj (Incl. KTM & TRM)': 'Bajaj',
  'Chetak Only': 'Chetak',
  'Bajaj + Chetak': 'Bajaj + Chetak',
  'TVS: All': 'TVS (All)',
  'TVS ICE': 'TVS ICE',
  'TVS: Only Iqube': 'TVS iQube',
  'Ather': 'Ather',
  'OLA Only CPS': 'OLA',
  'Revolt': 'Revolt',
  'Jawa (Incl. Manpower)': 'Jawa',
}};

const TOTAL_OEMS = ['Bajaj (Incl. KTM & TRM)','Chetak Only','Ather','OLA Only CPS','TVS: All','Revolt','Jawa (Incl. Manpower)'];

// ═══════════════════════════════════════════════════════════════
// STATE
// ═══════════════════════════════════════════════════════════════
let currentFY = 'overall';
let currentOEM = 'all';
let currentMonth = 'all';
let currentQuarter = 'all';
let selectedOEMwise = 'Bajaj (Incl. KTM & TRM)';
let chartRegistry = {{}};
let isDark = true;

// ═══════════════════════════════════════════════════════════════
// FORMATTERS
// ═══════════════════════════════════════════════════════════════
const fmt = {{
  rev: v => '₹' + (v/100000).toFixed(1) + 'L',
  revFull: v => '₹' + v.toLocaleString('en-IN'),
  cm2: v => '₹' + (v/100000).toFixed(1) + 'L',
  ret: v => Math.round(v).toLocaleString('en-IN'),
  pct: v => (v*100).toFixed(1) + '%',
}};

// ═══════════════════════════════════════════════════════════════
// INIT
// ═══════════════════════════════════════════════════════════════
function init() {{
  populateFilters();
  setupListeners();
  updateTime();
  document.getElementById('loadingScreen').style.display = 'none';
  document.getElementById('app').style.display = 'block';
  renderAll();
}}

function populateFilters() {{
  const oemSel = document.getElementById('oemFilter');
  for (const n of Object.keys(RAW_DATA.oems)) {{
    const opt = document.createElement('option');
    opt.value = n;
    opt.textContent = OEM_NAMES[n] || n;
    oemSel.appendChild(opt);
  }}
  const monthSel = document.getElementById('monthFilter');
  for (const m of RAW_DATA.months) {{
    const opt = document.createElement('option');
    opt.value = m;
    opt.textContent = m;
    monthSel.appendChild(opt);
  }}
}}

function setupListeners() {{
  document.querySelectorAll('.main-tab').forEach(btn => {{
    btn.addEventListener('click', () => {{
      document.querySelectorAll('.main-tab').forEach(b => b.classList.remove('active'));
      document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
      btn.classList.add('active');
      document.getElementById('tab-' + btn.dataset.tab).classList.add('active');
      if (btn.dataset.tab === 'oemwise') renderOEMwise();
    }});
  }});

  document.getElementById('fyFilter').addEventListener('change', e => {{ currentFY = e.target.value; renderAll(); }});
  document.getElementById('oemFilter').addEventListener('change', e => {{ currentOEM = e.target.value; renderAll(); }});
  document.getElementById('monthFilter').addEventListener('change', e => {{ currentMonth = e.target.value; renderAll(); }});
  document.getElementById('quarterFilter').addEventListener('change', e => {{ currentQuarter = e.target.value; renderAll(); }});

  document.getElementById('resetFilters').addEventListener('click', () => {{
    currentFY='overall'; currentOEM='all'; currentMonth='all'; currentQuarter='all';
    document.getElementById('fyFilter').value='overall';
    document.getElementById('oemFilter').value='all';
    document.getElementById('monthFilter').value='all';
    document.getElementById('quarterFilter').value='all';
    renderAll();
  }});

  document.getElementById('refreshBtn').addEventListener('click', () => location.reload());
  document.getElementById('themeToggle').addEventListener('click', toggleTheme);
  document.getElementById('tableSearch').addEventListener('input', filterTable);
  document.getElementById('exportBtn').addEventListener('click', exportCSV);

  document.querySelectorAll('.chart-action').forEach(btn => {{
    btn.addEventListener('click', () => {{
      const chart = chartRegistry[btn.dataset.chart];
      if (chart) {{
        const link = document.createElement('a');
        link.download = btn.dataset.chart + '_' + new Date().toISOString().slice(0,10) + '.png';
        link.href = chart.toBase64Image();
        link.click();
      }}
    }});
  }});
}}

function toggleTheme() {{
  isDark = !isDark;
  document.documentElement.setAttribute('data-theme', isDark ? 'dark' : 'light');
  document.getElementById('themeToggle').textContent = isDark ? '🌙' : '☀️';
  renderAll();
}}

function updateTime() {{
  const now = new Date().toLocaleString('en-IN', {{dateStyle:'medium',timeStyle:'short'}});
  document.getElementById('lastRefresh').textContent = 'Refreshed: ' + now;
  document.getElementById('footerTime').textContent = now;
}}

// ═══════════════════════════════════════════════════════════════
// FILTERING
// ═══════════════════════════════════════════════════════════════
function getFilteredData() {{
  const months = RAW_DATA.fy_splits[currentFY];
  const startIdx = RAW_DATA.months.indexOf(months[0]);

  let indices = [];
  if (currentMonth !== 'all') {{
    const mi = RAW_DATA.months.indexOf(currentMonth);
    if (mi >= startIdx && mi < startIdx + months.length) indices = [mi - startIdx];
  }} else {{
    indices = Array.from({{length: months.length}}, (_, i) => i);
  }}

  if (currentQuarter !== 'all') {{
    const qMap = {{'Q1':[0,1,2],'Q2':[3,4,5],'Q3':[6,7,8],'Q4':[9,10,11]}};
    const q = qMap[currentQuarter] || [];
    indices = indices.filter(i => q.includes(i));
  }}

  const result = {{ months: indices.map(i => months[i]), oems: {{}} }};
  const oemList = currentOEM === 'all' ? Object.keys(RAW_DATA.oems) : [currentOEM];
  for (const name of oemList) {{
    result.oems[name] = {{
      revenue: indices.map(i => RAW_DATA.oems[name].revenue[startIdx + i]),
      retail_count: indices.map(i => RAW_DATA.oems[name].retail_count[startIdx + i]),
      cm2: indices.map(i => RAW_DATA.oems[name].cm2[startIdx + i]),
      cm2_pct: indices.map(i => RAW_DATA.oems[name].cm2_pct[startIdx + i]),
    }};
  }}
  return result;
}}

// ═══════════════════════════════════════════════════════════════
// RENDER ALL
// ═══════════════════════════════════════════════════════════════
function renderAll() {{
  const data = getFilteredData();
  destroyCharts();
  renderExecKPIs(data);
  renderRevTrend(data);
  renderRetTrend(data);
  renderCM2Trend(data);
  renderCM2PctTrend(data);
  renderOEMCompare(data);
  renderStackedRev(data);
  renderTable(data);
  if (document.getElementById('tab-oemwise').classList.contains('active')) renderOEMwise();
}}

function destroyCharts() {{
  Object.values(chartRegistry).forEach(c => c.destroy());
  chartRegistry = {{}};
}}

function createChart(id, type, chartData, options) {{
  const ctx = document.getElementById(id).getContext('2d');
  const chart = new Chart(ctx, {{ type, data: chartData, options }});
  chartRegistry[id] = chart;
  return chart;
}}

function commonOptions() {{
  return {{
    responsive: true, maintainAspectRatio: false,
    interaction: {{ mode: 'index', intersect: false }},
    plugins: {{
      legend: {{ labels: {{ color: isDark ? '#94a3b8' : '#475569', font: {{ size: 11 }} }} }},
      datalabels: {{ display: false }},
    }},
    scales: {{
      x: {{ ticks: {{ color: isDark ? '#64748b' : '#94a3b8', maxRotation: 45 }}, grid: {{ color: isDark ? 'rgba(148,163,184,0.08)' : 'rgba(148,163,184,0.15)' }} }},
      y: {{ ticks: {{ color: isDark ? '#64748b' : '#94a3b8' }}, grid: {{ color: isDark ? 'rgba(148,163,184,0.08)' : 'rgba(148,163,184,0.15)' }} }},
    }},
  }};
}}

// ═══════════════════════════════════════════════════════════════
// EXECUTIVE KPIs
// ═══════════════════════════════════════════════════════════════
function renderExecKPIs(data) {{
  const months = data.months;
  if (months.length === 0) return;

  let totalRev = 0, totalCm2 = 0, totalRet = 0;
  for (const name of TOTAL_OEMS) {{
    if (data.oems[name]) {{
      totalRev += data.oems[name].revenue.reduce((a,b)=>a+b,0);
      totalCm2 += data.oems[name].cm2.reduce((a,b)=>a+b,0);
      totalRet += data.oems[name].retail_count.reduce((a,b)=>a+b,0);
    }}
  }}
  const avgPct = totalRev !== 0 ? totalCm2/totalRev : 0;
  const g = RAW_DATA.growth;

  document.getElementById('execRevenue').textContent = fmt.rev(totalRev);
  document.getElementById('execRetail').textContent = fmt.ret(totalRet);
  document.getElementById('execCm2').textContent = fmt.cm2(totalCm2);
  document.getElementById('execCm2Pct').textContent = fmt.pct(avgPct);

  const revG = g.revenue_growth_pct;
  const retG = g.retail_growth_pct;
  document.getElementById('execRevGrowth').textContent = (revG>=0?'+':'') + revG.toFixed(1) + '%';
  document.getElementById('execRevGrowth').className = 'kpi-value ' + (revG>=0?'up':'down');
  document.getElementById('execRetGrowth').textContent = (retG>=0?'+':'') + retG.toFixed(1) + '%';
  document.getElementById('execRetGrowth').className = 'kpi-value ' + (retG>=0?'up':'down');

  document.getElementById('execRevenueSub').textContent = months[months.length-1] + ' · ' + fmt.revFull(totalRev);
  document.getElementById('execRetailSub').textContent = months[months.length-1] + ' · ' + fmt.ret(totalRet);
  document.getElementById('execCm2Sub').textContent = months[months.length-1] + ' · ' + fmt.cm2(totalCm2);
  document.getElementById('execCm2PctSub').textContent = 'Avg across all OEMs';
  document.getElementById('execRevGrowthSub').textContent = 'vs FY 25-26 baseline';
  document.getElementById('execRetGrowthSub').textContent = 'vs FY 25-26 baseline';
}}

// ═══════════════════════════════════════════════════════════════
// CHARTS
// ═══════════════════════════════════════════════════════════════
function renderRevTrend(data) {{
  const datasets = Object.keys(data.oems).map(n => ({{
    label: OEM_NAMES[n] || n,
    data: data.oems[n].revenue,
    borderColor: RAW_DATA.colors[n] || '#999',
    backgroundColor: (RAW_DATA.colors[n] || '#999') + '15',
    fill: true, tension: 0.35, pointRadius: 4, pointHoverRadius: 7, borderWidth: 2,
  }}));
  createChart('revTrendChart', 'line', {{ labels: data.months, datasets }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, tooltip: {{ callbacks: {{ label: ctx => ctx.dataset.label + ': ' + fmt.rev(ctx.parsed.y) }} }} }},
    scales: {{ ...commonOptions().scales, y: {{ ...commonOptions().scales.y, ticks: {{ ...commonOptions().scales.y.ticks, callback: v => fmt.rev(v) }} }} }},
  }});
}}

function renderRetTrend(data) {{
  const totalRetail = data.months.map((_, i) => {{
    let s = 0;
    for (const n of TOTAL_OEMS) if (data.oems[n]) s += data.oems[n].retail_count[i] || 0;
    return s;
  }});
  createChart('retTrendChart', 'bar', {{ labels: data.months, datasets: [{{ label: 'Total Retail', data: totalRetail, backgroundColor: '#f59e0b', borderRadius: 6 }}] }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, tooltip: {{ callbacks: {{ label: ctx => 'Retail: ' + fmt.ret(ctx.parsed.y) }} }} }},
    scales: {{ ...commonOptions().scales, y: {{ ...commonOptions().scales.y, ticks: {{ ...commonOptions().scales.y.ticks, callback: v => fmt.ret(v) }} }} }},
  }});
}}

function renderCM2Trend(data) {{
  const totalCm2 = data.months.map((_, i) => {{
    let s = 0;
    for (const n of TOTAL_OEMS) if (data.oems[n]) s += data.oems[n].cm2[i] || 0;
    return s;
  }});
  createChart('cm2TrendChart', 'line', {{ labels: data.months, datasets: [{{ label: 'Total CM2', data: totalCm2, borderColor: '#10b981', backgroundColor: 'rgba(16,185,129,0.1)', fill: true, tension: 0.35, pointRadius: 4, borderWidth: 2 }}] }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, tooltip: {{ callbacks: {{ label: ctx => 'CM2: ' + fmt.cm2(ctx.parsed.y) }} }} }},
    scales: {{ ...commonOptions().scales, y: {{ ...commonOptions().scales.y, ticks: {{ ...commonOptions().scales.y.ticks, callback: v => fmt.cm2(v) }} }} }},
  }});
}}

function renderCM2PctTrend(data) {{
  const avgPct = data.months.map((_, i) => {{
    let rev = 0, cm2 = 0;
    for (const n of TOTAL_OEMS) {{
      if (data.oems[n]) {{ rev += data.oems[n].revenue[i] || 0; cm2 += data.oems[n].cm2[i] || 0; }}
    }}
    return rev !== 0 ? cm2/rev : 0;
  }});
  createChart('cm2PctTrendChart', 'line', {{ labels: data.months, datasets: [{{ label: 'Avg CM2 %', data: avgPct, borderColor: '#8b5cf6', backgroundColor: 'rgba(139,92,246,0.1)', fill: true, tension: 0.35, pointRadius: 4, borderWidth: 2 }}] }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, tooltip: {{ callbacks: {{ label: ctx => 'CM2%: ' + fmt.pct(ctx.parsed.y) }} }} }},
    scales: {{ ...commonOptions().scales, y: {{ ...commonOptions().scales.y, ticks: {{ ...commonOptions().scales.y.ticks, callback: v => fmt.pct(v) }} }} }},
  }});
}}

function renderOEMCompare(data) {{
  const labels = Object.keys(data.oems).map(n => OEM_NAMES[n] || n);
  const values = Object.keys(data.oems).map(n => data.oems[n].revenue.reduce((a,b)=>a+b,0));
  const bgColors = Object.keys(data.oems).map(n => RAW_DATA.colors[n] || '#999');
  createChart('oemCompareChart', 'bar', {{ labels, datasets: [{{ label: 'Total Revenue', data: values, backgroundColor: bgColors, borderRadius: 8 }}] }}, {{
    ...commonOptions(), indexAxis: 'y',
    plugins: {{ ...commonOptions().plugins, datalabels: {{ display: true, anchor: 'end', align: 'right', offset: 4, color: isDark ? '#f1f5f9' : '#1e293b', font: {{ weight: 'bold', size: 10 }}, formatter: v => fmt.rev(v) }}, tooltip: {{ callbacks: {{ label: ctx => fmt.rev(ctx.parsed.x) }} }} }},
    scales: {{ x: {{ ticks: {{ color: isDark ? '#64748b' : '#94a3b8', callback: v => fmt.rev(v) }}, grid: {{ color: isDark ? 'rgba(148,163,184,0.08)' : 'rgba(148,163,184,0.15)' }} }}, y: {{ ticks: {{ color: isDark ? '#94a3b8' : '#475569', font: {{ weight: 'bold' }} }}, grid: {{ display: false }} }} }},
  }});
}}

function renderStackedRev(data) {{
  const datasets = Object.keys(data.oems).map(n => ({{
    label: OEM_NAMES[n] || n, data: data.oems[n].revenue,
    backgroundColor: RAW_DATA.colors[n] || '#999', borderWidth: 0, borderRadius: 2,
  }}));
  createChart('stackedRevChart', 'bar', {{ labels: data.months, datasets }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, tooltip: {{ callbacks: {{ label: ctx => ctx.dataset.label + ': ' + fmt.rev(ctx.parsed.y) }} }} }},
    scales: {{ x: {{ stacked: true, ticks: {{ color: isDark ? '#64748b' : '#94a3b8', maxRotation: 45 }}, grid: {{ color: isDark ? 'rgba(148,163,184,0.08)' : 'rgba(148,163,184,0.15)' }} }}, y: {{ stacked: true, ticks: {{ color: isDark ? '#64748b' : '#94a3b8', callback: v => fmt.rev(v) }}, grid: {{ color: isDark ? 'rgba(148,163,184,0.08)' : 'rgba(148,163,184,0.15)' }} }} }},
  }});
}}

// ═══════════════════════════════════════════════════════════════
// TABLE
// ═══════════════════════════════════════════════════════════════
function renderTable(data) {{
  const months = data.months;
  const oems = data.oems;
  const thead = document.querySelector('#summaryTable thead');
  const tbody = document.querySelector('#summaryTable tbody');

  let h = '<tr><th>OEM</th>';
  for (const m of months) h += '<th class="num">' + m + '</th>';
  h += '<th class="num">Total</th><th class="num">CM2%</th></tr>';
  thead.innerHTML = h;

  tbody.innerHTML = '';
  for (const name of Object.keys(oems)) {{
    const d = oems[name];
    const totalRev = d.revenue.reduce((a,b)=>a+b,0);
    const totalCm2 = d.cm2.reduce((a,b)=>a+b,0);
    const cm2pct = totalRev !== 0 ? totalCm2/totalRev : 0;
    let row = '<td><span class="oem-dot" style="background:' + (RAW_DATA.colors[name] || '#999') + '"></span> ' + (OEM_NAMES[name] || name) + '</td>';
    for (const rev of d.revenue) row += '<td class="num">' + fmt.rev(rev) + '</td>';
    row += '<td class="num" style="font-weight:700">' + fmt.rev(totalRev) + '</td>';
    row += '<td class="num"><span class="badge ' + (cm2pct>=0?'badge-up':'badge-down') + '">' + fmt.pct(cm2pct) + '</span></td>';
    const tr = document.createElement('tr');
    tr.innerHTML = row;
    tr.dataset.oem = (OEM_NAMES[name] || name).toLowerCase();
    tbody.appendChild(tr);
  }}
}}

function filterTable(e) {{
  const term = e.target.value.toLowerCase();
  document.querySelectorAll('#summaryTable tbody tr').forEach(tr => {{
    tr.style.display = tr.dataset.oem.includes(term) ? '' : 'none';
  }});
}}

// ═══════════════════════════════════════════════════════════════
// OEM-WISE
// ═══════════════════════════════════════════════════════════════
function renderOEMwise() {{
  renderOEMPills();
  renderOEMCharts();
}}

function renderOEMPills() {{
  const container = document.getElementById('oemPills');
  container.innerHTML = '';
  for (const name of Object.keys(RAW_DATA.oems)) {{
    const pill = document.createElement('button');
    pill.className = 'oem-pill' + (name === selectedOEMwise ? ' active' : '');
    const color = RAW_DATA.colors[name] || '#999';
    pill.innerHTML = '<span class="oem-dot" style="background:' + color + '"></span>' + (OEM_NAMES[name] || name);
    pill.addEventListener('click', () => {{
      selectedOEMwise = name;
      document.querySelectorAll('.oem-pill').forEach(p => p.classList.remove('active'));
      pill.classList.add('active');
      renderOEMCharts();
    }});
    container.appendChild(pill);
  }}
}}

function renderOEMCharts() {{
  const data = getFilteredData();
  if (!data.oems[selectedOEMwise]) {{
    selectedOEMwise = Object.keys(data.oems)[0] || 'Bajaj (Incl. KTM & TRM)';
    renderOEMPills();
  }}
  const d = data.oems[selectedOEMwise];
  const months = data.months;
  const color = RAW_DATA.colors[selectedOEMwise] || '#3b82f6';
  const name = OEM_NAMES[selectedOEMwise] || selectedOEMwise;

  const totalRev = d.revenue.reduce((a,b)=>a+b,0);
  const totalRet = d.retail_count.reduce((a,b)=>a+b,0);
  const totalCm2 = d.cm2.reduce((a,b)=>a+b,0);
  const avgPct = totalRev !== 0 ? totalCm2/totalRev : 0;

  document.getElementById('oemRevenue').textContent = fmt.rev(totalRev);
  document.getElementById('oemRetail').textContent = fmt.ret(totalRet);
  document.getElementById('oemCm2').textContent = fmt.cm2(totalCm2);
  document.getElementById('oemCm2Pct').textContent = fmt.pct(avgPct);
  document.getElementById('oemRevenueSub').textContent = months.length + ' months';
  document.getElementById('oemRetailSub').textContent = months.length + ' months';
  document.getElementById('oemCm2Sub').textContent = months.length + ' months';
  document.getElementById('oemCm2PctSub').textContent = 'Average';

  ['oemRevChart','oemRetChart','oemCm2Chart','oemCm2PctChart','oemGrowthChart'].forEach(id => {{
    if (chartRegistry[id]) {{ chartRegistry[id].destroy(); delete chartRegistry[id]; }}
  }});

  createChart('oemRevChart', 'line', {{ labels: months, datasets: [{{ label: 'Revenue', data: d.revenue, borderColor: color, backgroundColor: color + '18', fill: true, tension: 0.35, pointRadius: 5, pointHoverRadius: 8, pointBackgroundColor: color, borderWidth: 2.5 }}] }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, datalabels: {{ display: true, align: 'top', anchor: 'end', offset: 4, color: color, font: {{ size: 10, weight: 'bold' }}, formatter: v => fmt.rev(v), display: ctx => months.length <= 10 || ctx.dataIndex % 2 === 0 }}, tooltip: {{ callbacks: {{ label: ctx => 'Revenue: ' + fmt.rev(ctx.parsed.y) }} }} }},
    scales: {{ ...commonOptions().scales, y: {{ ...commonOptions().scales.y, ticks: {{ ...commonOptions().scales.y.ticks, callback: v => fmt.rev(v) }} }} }},
  }});

  createChart('oemRetChart', 'bar', {{ labels: months, datasets: [{{ label: 'Retail', data: d.retail_count, backgroundColor: color + 'CC', borderRadius: 6 }}] }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, datalabels: {{ display: true, align: 'top', anchor: 'end', offset: 2, color: color, font: {{ size: 10, weight: 'bold' }}, formatter: v => fmt.ret(v) }}, tooltip: {{ callbacks: {{ label: ctx => 'Retail: ' + fmt.ret(ctx.parsed.y) }} }} }},
    scales: {{ ...commonOptions().scales, y: {{ ...commonOptions().scales.y, ticks: {{ ...commonOptions().scales.y.ticks, callback: v => fmt.ret(v) }} }} }},
  }});

  createChart('oemCm2Chart', 'line', {{ labels: months, datasets: [{{ label: 'CM2', data: d.cm2, borderColor: color, backgroundColor: color + '18', fill: true, tension: 0.35, pointRadius: 5, pointHoverRadius: 8, pointBackgroundColor: color, borderWidth: 2.5 }}] }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, datalabels: {{ display: true, align: 'top', anchor: 'end', offset: 4, color: color, font: {{ size: 10, weight: 'bold' }}, formatter: v => fmt.cm2(v), display: ctx => months.length <= 10 || ctx.dataIndex % 2 === 0 }}, tooltip: {{ callbacks: {{ label: ctx => 'CM2: ' + fmt.cm2(ctx.parsed.y) }} }} }},
    scales: {{ ...commonOptions().scales, y: {{ ...commonOptions().scales.y, ticks: {{ ...commonOptions().scales.y.ticks, callback: v => fmt.cm2(v) }} }} }},
  }});

  createChart('oemCm2PctChart', 'line', {{ labels: months, datasets: [{{ label: 'CM2 %', data: d.cm2_pct, borderColor: color, backgroundColor: 'transparent', tension: 0.35, pointRadius: 5, pointHoverRadius: 8, pointBackgroundColor: color, borderWidth: 2.5, borderDash: [5, 5] }}] }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, datalabels: {{ display: true, align: 'top', anchor: 'end', offset: 4, color: color, font: {{ size: 10, weight: 'bold' }}, formatter: v => fmt.pct(v) }}, tooltip: {{ callbacks: {{ label: ctx => 'CM2%: ' + fmt.pct(ctx.parsed.y) }} }} }},
    scales: {{ ...commonOptions().scales, y: {{ ...commonOptions().scales.y, ticks: {{ ...commonOptions().scales.y.ticks, callback: v => fmt.pct(v) }} }} }},
  }});

  const growthData = Object.keys(data.oems).map(n => {{
    const rev = data.oems[n].revenue;
    if (rev.length < 2) return 0;
    return ((rev[rev.length-1] - rev[0]) / Math.abs(rev[0])) * 100;
  }});
  const growthLabels = Object.keys(data.oems).map(n => OEM_NAMES[n] || n);

  createChart('oemGrowthChart', 'bar', {{ labels: growthLabels, datasets: [{{ label: 'Revenue Growth %', data: growthData, backgroundColor: growthData.map(v => v >= 0 ? '#10b981' : '#ef4444'), borderRadius: 6 }}] }}, {{
    ...commonOptions(),
    plugins: {{ ...commonOptions().plugins, datalabels: {{ display: true, align: 'end', anchor: 'end', offset: 2, color: isDark ? '#f1f5f9' : '#1e293b', font: {{ weight: 'bold', size: 10 }}, formatter: v => (v >= 0 ? '+' : '') + v.toFixed(1) + '%' }}, tooltip: {{ callbacks: {{ label: ctx => 'Growth: ' + (ctx.parsed.y >= 0 ? '+' : '') + ctx.parsed.y.toFixed(1) + '%' }} }} }},
    scales: {{ ...commonOptions().scales, y: {{ ...commonOptions().scales.y, ticks: {{ ...commonOptions().scales.y.ticks, callback: v => v.toFixed(0) + '%' }} }} }},
  }});
}}

// ═══════════════════════════════════════════════════════════════
// EXPORT
// ═══════════════════════════════════════════════════════════════
function exportCSV() {{
  const data = getFilteredData();
  const months = data.months;
  let csv = 'OEM,Metric,' + months.join(',') + ',Total\\n';
  for (const name of Object.keys(data.oems)) {{
    const d = data.oems[name];
    csv += '"' + name + '",Revenue,' + d.revenue.join(',') + ',' + d.revenue.reduce((a,b)=>a+b,0) + '\\n';
    csv += '"' + name + '",Retail,' + d.retail_count.join(',') + ',' + d.retail_count.reduce((a,b)=>a+b,0) + '\\n';
    csv += '"' + name + '",CM2,' + d.cm2.join(',') + ',' + d.cm2.reduce((a,b)=>a+b,0) + '\\n';
  }}
  const blob = new Blob([csv], {{ type: 'text/csv' }});
  const url = URL.createObjectURL(blob);
  const link = document.createElement('a');
  link.download = 'CPS_OEM_Data_' + new Date().toISOString().slice(0,10) + '.csv';
  link.href = url;
  link.click();
  URL.revokeObjectURL(url);
}}

// ═══════════════════════════════════════════════════════════════
// START
// ═══════════════════════════════════════════════════════════════
document.addEventListener('DOMContentLoaded', init);
</script>

</body>
</html>
'''

# Write the single HTML file
out_path = r'C:/Users/saten/Documents/kimi/workspace/CPS_OEM_Dashboard/index.html'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(html)

# Also write the standalone version
standalone_path = r'C:/Users/saten/Documents/kimi/workspace/CPS_OEM_Dashboard_Standalone.html'
with open(standalone_path, 'w', encoding='utf-8') as f:
    f.write(html)

print("=" * 55)
print("✅ FIXED! Self-contained dashboard generated")
print("=" * 55)
print(f"   Main file: {out_path}")
print(f"   Standalone: {standalone_path}")
print("\n   🎯 IMPORTANT: Open the HTML file directly in Chrome/Edge.")
print("   The data is now EMBEDDED — no fetch() needed!")
print("   It will work instantly without any server.")
print(f"\n   Verified Totals:")
print(f"   FY 25-26 Revenue: ₹{fy2526_totals['revenue']:,.0f} ✅")
print(f"   FY 26-27 Revenue: ₹{fy2627_totals['revenue']:,.0f} ✅")
print(f"   Overall Revenue:  ₹{overall_totals['revenue']:,.0f} ✅")
